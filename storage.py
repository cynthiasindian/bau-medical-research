# =============================================================================
#  storage.py  ·  the spreadsheet read/write layer
#
#  One Storage class hides whether data lives in a local Excel file or a Google
#  Sheet. The rest of the app only calls the high-level methods:
#     list_records(), get_patient(), create_patient(), save_section(),
#     delete_patient()
#
#  IMPORTANT — concurrency / data safety:
#  Every write is a read-modify-write of the WHOLE sheet, performed under a file
#  lock (Excel) so two students saving at once cannot clobber each other's row.
#  This is simple and reliable for a small team; it is not built for hundreds of
#  simultaneous writers.
# =============================================================================
import os
from pathlib import Path

import pandas as pd
from filelock import FileLock

import core
import surveys


# -----------------------------------------------------------------------------
# Shared grouped-header schema (used by BOTH backends):
#   row 1: merged survey banners   PATIENT | PRE-SURGERY | POST-SURGERY
#   row 2: merged section titles   (one cell per section)
#   row 3: the full question text  (as it appears in the survey)
#   row 4: field headers           (the canonical column keys — what is read)
#   row 5+: one patient per row
# -----------------------------------------------------------------------------
HEADER_ROWS = 4

# banner -> fill color (hex, no #)
_FILLS = {"PATIENT": "D9D9D9", "PRE-SURGERY": "C6E0B4",
          "POST-SURGERY": "BDD7EE", "": "EEEEEE"}


def _schema_blocks(df_columns):
    """Yield (banner, section_title, [columns]) column groups in order."""
    blocks = [("PATIENT", "Record info", list(core.META_COLS))]
    for survey, banner in (("pre", "PRE-SURGERY"), ("post", "POST-SURGERY")):
        for s in surveys.sections_for(survey):
            sec_cols = [core.status_col(s["key"])] + [f["key"] for f in s["fields"]]
            blocks.append((banner, f"Section {s['number']} · {s['title']}", sec_cols))
    known = {c for _, _, cols in blocks for c in cols}
    extras = [c for c in df_columns if c not in known]
    if extras:
        blocks.append(("", "Other", extras))
    return [(b, t, [c for c in cols if c in df_columns])
            for b, t, cols in blocks if any(c in df_columns for c in cols)]


def _column_labels():
    """key -> full question text, for the human-readable header row."""
    labels = {"patient_id": "Patient ID (auto)",
              "display_id": "Hospital ID (shown in lists)",
              "created_at": "Record created at",
              "updated_at": "Last updated at"}
    for s in surveys.SECTIONS:
        labels[core.status_col(s["key"])] = "Section completion status"
        for f in s["fields"]:
            labels[f["key"]] = f["label"]
    return labels


# -----------------------------------------------------------------------------
# Backends: each only needs to read the whole sheet into a DataFrame and write
# the whole DataFrame back. All business logic lives in the Storage facade.
# -----------------------------------------------------------------------------
class _ExcelBackend:
    """Local .xlsx workbook with the grouped 4-row header (see above).
    Reads use row 4 as the header, so rows 1-3 are presentation only."""

    HEADER_ROWS = HEADER_ROWS

    def __init__(self, cfg):
        self.path = Path(cfg["excel"]["path"])
        self.path.parent.mkdir(parents=True, exist_ok=True)
        # Lock file sits next to the workbook.
        self.lock = FileLock(str(self.path) + ".lock")

    def read_df(self):
        if not self.path.exists():
            return pd.DataFrame(columns=core.COLUMNS)
        # keep_default_na=False -> empty cells become "" (not NaN), so values
        # round-trip cleanly; numbers still read back as numbers.
        # Try the current layout first, then older layouts (3-row grouped
        # header, flat single header) — the next write migrates the file.
        for header_row in (self.HEADER_ROWS - 1, 2, 0):
            try:
                df = pd.read_excel(self.path, header=header_row,
                                   dtype=object, keep_default_na=False)
            except ValueError:  # fewer rows than this header layout needs
                continue
            if "patient_id" in df.columns:
                return df
        # Last resort: flat read, even if the id column is missing.
        return pd.read_excel(self.path, dtype=object, keep_default_na=False)

    def write_df(self, df):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "patients"

        blocks = _schema_blocks(list(df.columns))
        df = df[[c for _, _, cols in blocks for c in cols]]

        fills = _FILLS
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold = Font(bold=True)

        # --- row 2: one merged cell per section --------------------------------
        col, row1_groups = 1, []
        for banner, title, cols in blocks:
            end = col + len(cols) - 1
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=end)
            c = ws.cell(row=2, column=col, value=title)
            c.alignment, c.font = center, bold
            c.fill = PatternFill(fill_type="solid", fgColor=fills.get(banner, "EEEEEE"))
            row1_groups.append((banner, col, end))
            col = end + 1

        # --- row 1: merge consecutive sections of the same survey --------------
        i = 0
        while i < len(row1_groups):
            banner, start, end = row1_groups[i]
            while i + 1 < len(row1_groups) and row1_groups[i + 1][0] == banner:
                i += 1
                end = row1_groups[i][2]
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            c = ws.cell(row=1, column=start, value=banner)
            c.alignment, c.font = center, Font(bold=True, size=12)
            c.fill = PatternFill(fill_type="solid", fgColor=fills.get(banner, "EEEEEE"))
            i += 1

        # --- row 3: the full question text, as it appears in the survey --------
        labels = _column_labels()
        wrap_top = Alignment(vertical="top", wrap_text=True)
        for ci, name in enumerate(df.columns, start=1):
            c = ws.cell(row=3, column=ci, value=labels.get(name, ""))
            c.alignment = wrap_top
            c.font = Font(italic=True, size=9)
        ws.row_dimensions[3].height = 90

        # --- row 4: the canonical field headers (what the app reads) -----------
        for ci, name in enumerate(df.columns, start=1):
            c = ws.cell(row=self.HEADER_ROWS, column=ci, value=name)
            c.font = bold
            ws.column_dimensions[get_column_letter(ci)].width = \
                max(12, min(28, len(str(name)) + 2))

        # Keep headers + patient id/name visible while scrolling.
        ws.freeze_panes = "C5"

        # --- data, one patient per row from row 5 ------------------------------
        for record in df.itertuples(index=False):
            ws.append(list(record))
        wb.save(self.path)


class _GoogleSheetsBackend:
    """Google Sheet with the same grouped 4-row header as the Excel backend
    (merged banners/sections, question text, keys, data from row 5)."""

    HEADER_ROWS = HEADER_ROWS

    def __init__(self, cfg):
        import gspread  # imported lazily so Excel users need not install it
        gs = cfg["google_sheets"]
        key_file = Path(gs["service_account_file"])
        if key_file.exists():
            self.client = gspread.service_account(filename=key_file)
        else:
            # Deployed (e.g. Streamlit Community Cloud): the key file is not
            # in the repo, so read the credentials from st.secrets instead.
            import streamlit as st
            self.client = gspread.service_account_from_dict(
                dict(st.secrets["gcp_service_account"]))
        self.sh = self.client.open_by_key(gs["spreadsheet_id"])
        self.ws_name = gs["worksheet_name"]
        # gspread has its own server-side consistency; a no-op local lock keeps
        # the Storage code identical across backends.
        from contextlib import nullcontext
        self.lock = nullcontext()

    def _worksheet(self):
        try:
            return self.sh.worksheet(self.ws_name)
        except Exception:
            return self.sh.add_worksheet(self.ws_name, rows=100, cols=max(26, len(core.COLUMNS)))

    def read_df(self):
        ws = self._worksheet()
        rows = ws.get_all_values()
        # Current layout: keys in row 4. Legacy flat layout: keys in row 1.
        header_idx = next((i for i in (self.HEADER_ROWS - 1, 0)
                           if len(rows) > i and "patient_id" in rows[i]), None)
        if header_idx is None:
            return pd.DataFrame(columns=core.COLUMNS)
        header = rows[header_idx]
        width = len(header)
        data = [r[:width] + [""] * (width - len(r)) for r in rows[header_idx + 1:]]
        return pd.DataFrame(data, columns=header, dtype=object)

    def write_df(self, df):
        ws = self._worksheet()
        sid = ws.id

        blocks = _schema_blocks(list(df.columns))
        df = df[[c for _, _, cols in blocks for c in cols]]
        ncols = len(df.columns)

        # ---- build the 4 header rows + data ----------------------------------
        row1, row2 = [""] * ncols, [""] * ncols
        sections, col = [], 0                    # (banner, start, end) 0-based
        for banner, title, cols in blocks:
            end = col + len(cols) - 1
            row2[col] = title
            sections.append((banner, col, end))
            col = end + 1
        banners, i = [], 0                       # consecutive same-banner spans
        while i < len(sections):
            banner, start, end = sections[i]
            while i + 1 < len(sections) and sections[i + 1][0] == banner:
                i += 1
                end = sections[i][2]
            banners.append((banner, start, end))
            row1[start] = banner
            i += 1
        labels = _column_labels()
        row3 = [labels.get(c, "") for c in df.columns]
        row4 = list(df.columns)
        clean = df.astype(object).where(pd.notna(df), "")
        data = [[v.item() if hasattr(v, "item") else v for v in r]
                for r in clean.values.tolist()]
        values = [row1, row2, row3, row4] + data

        # ---- write values (old merges must go first, or re-merging errors) ---
        self.sh.batch_update({"requests": [
            {"unmergeCells": {"range": {"sheetId": sid}}}]})
        ws.resize(rows=max(50, len(values) + 20), cols=ncols)
        ws.clear()
        ws.update(values, value_input_option="RAW")

        # ---- merges + styling, mirroring the Excel layout --------------------
        def rng(row, c1, c2):
            return {"sheetId": sid, "startRowIndex": row, "endRowIndex": row + 1,
                    "startColumnIndex": c1, "endColumnIndex": c2 + 1}

        def color(hexs):
            return {"red": int(hexs[0:2], 16) / 255,
                    "green": int(hexs[2:4], 16) / 255,
                    "blue": int(hexs[4:6], 16) / 255}

        def fmt(range_, fields, **props):
            return {"repeatCell": {"range": range_,
                                   "cell": {"userEnteredFormat": props},
                                   "fields": "userEnteredFormat(" + fields + ")"}}

        FROZEN_COLS = 2

        def add_merge(reqs, row, start, end):
            # Sheets refuses frozen columns that cut through a merged cell,
            # so split any merge straddling the freeze line into two.
            parts = ([(start, FROZEN_COLS - 1), (FROZEN_COLS, end)]
                     if start < FROZEN_COLS <= end else [(start, end)])
            for s, e in parts:
                if e > s:
                    reqs.append({"mergeCells": {"range": rng(row, s, e),
                                                "mergeType": "MERGE_ALL"}})

        reqs = []
        for banner, start, end in banners:
            add_merge(reqs, 0, start, end)
            reqs.append(fmt(rng(0, start, end),
                            "backgroundColor,textFormat,horizontalAlignment,verticalAlignment",
                            backgroundColor=color(_FILLS.get(banner, "EEEEEE")),
                            textFormat={"bold": True, "fontSize": 12},
                            horizontalAlignment="CENTER", verticalAlignment="MIDDLE"))
        for banner, start, end in sections:
            add_merge(reqs, 1, start, end)
            reqs.append(fmt(rng(1, start, end),
                            "backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy",
                            backgroundColor=color(_FILLS.get(banner, "EEEEEE")),
                            textFormat={"bold": True},
                            horizontalAlignment="CENTER", verticalAlignment="MIDDLE",
                            wrapStrategy="WRAP"))
        reqs.append(fmt(rng(2, 0, ncols - 1), "textFormat,wrapStrategy,verticalAlignment",
                        textFormat={"italic": True, "fontSize": 8},
                        wrapStrategy="WRAP", verticalAlignment="TOP"))
        reqs.append(fmt(rng(3, 0, ncols - 1), "textFormat",
                        textFormat={"bold": True}))
        reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid,
                           "gridProperties": {"frozenRowCount": self.HEADER_ROWS,
                                              "frozenColumnCount": 2}},
            "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"}})
        reqs.append({"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "ROWS",
                      "startIndex": 2, "endIndex": 3},
            "properties": {"pixelSize": 110}, "fields": "pixelSize"}})
        reqs.append({"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "COLUMNS",
                      "startIndex": 0, "endIndex": ncols},
            "properties": {"pixelSize": 140}, "fields": "pixelSize"}})
        self.sh.batch_update({"requests": reqs})


# -----------------------------------------------------------------------------
# Storage facade
# -----------------------------------------------------------------------------
class Storage:
    def __init__(self, config):
        backend = config["storage"]["backend"]
        if backend == "excel":
            self._b = _ExcelBackend(config)
        elif backend == "google_sheets":
            self._b = _GoogleSheetsBackend(config)
        else:
            raise ValueError(f"Unknown storage backend: {backend!r}")
        self.backend_name = backend

    # ---- schema helpers -----------------------------------------------------
    def _read_normalized(self):
        """Read sheet and guarantee every expected column exists, in order."""
        df = self._b.read_df()
        for col in core.COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df.where(pd.notna(df), "")  # NaN -> "" without dtype downcasting
        # Unknown columns (e.g. from removed survey fields) are kept only if
        # they still hold any data, so retired placeholders fade out safely.
        extras = [c for c in df.columns if c not in core.COLUMNS
                  and (df[c].astype(str).str.strip() != "").any()]
        return df[core.COLUMNS + extras]

    # ---- read API -----------------------------------------------------------
    def list_records(self):
        """Return list of dict records (one per patient)."""
        df = self._read_normalized()
        return df.to_dict(orient="records")

    def get_patient(self, patient_id):
        df = self._read_normalized()
        match = df[df["patient_id"].astype(str) == str(patient_id)]
        if match.empty:
            return None
        return match.iloc[0].to_dict()

    # ---- write API ----------------------------------------------------------
    def create_patient(self, demo_cells):
        """Create a new patient row from Demographics (Pre §1) cell values.

        Returns the generated patient_id.
        """
        with self._b.lock:
            df = self._read_normalized()
            pid = core.next_patient_id(df["patient_id"].tolist())

            row = core.empty_record()
            row.update(demo_cells)
            row["patient_id"] = pid
            row["display_id"] = str(demo_cells.get(surveys.DISPLAY_ID_FIELD, "")).strip()
            row["created_at"] = core.now_stamp()
            row["updated_at"] = core.now_stamp()
            # Compute Demographics completion from the just-entered values.
            row[core.status_col("pre_s1")] = core.section_status(
                surveys.get_section("pre_s1"), row)

            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
            self._b.write_df(df)
        return pid

    def save_section(self, patient_id, section_key, cells):
        """Update one section's cells for an existing patient and refresh its
        completion status. Only the given section's columns are touched, so the
        rest of the pre/post data on that row is preserved."""
        section = surveys.get_section(section_key)
        with self._b.lock:
            df = self._read_normalized()
            mask = df["patient_id"].astype(str) == str(patient_id)
            if not mask.any():
                raise KeyError(f"Patient not found: {patient_id}")
            idx = df.index[mask][0]

            for key, val in cells.items():
                df.at[idx, key] = val

            # Recompute completion status from the merged row values.
            merged = df.loc[idx].to_dict()
            df.at[idx, core.status_col(section_key)] = core.section_status(section, merged)

            # Keep display_id in sync if the name was edited in Demographics.
            if section_key == "pre_s1":
                df.at[idx, "display_id"] = str(
                    cells.get(surveys.DISPLAY_ID_FIELD, df.at[idx, "display_id"])).strip()

            df.at[idx, "updated_at"] = core.now_stamp()
            self._b.write_df(df)

    def delete_patient(self, patient_id):
        """Permanently remove a patient's row (all pre + post answers)."""
        with self._b.lock:
            df = self._read_normalized()
            mask = df["patient_id"].astype(str) == str(patient_id)
            if not mask.any():
                raise KeyError(f"Patient not found: {patient_id}")
            self._b.write_df(df[~mask].reset_index(drop=True))
